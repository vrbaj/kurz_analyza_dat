from openpyxl import Workbook, load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import Border, Side



def nastav_ohraniceni(aktivni_sesit, rozsah_bunek):
    styl_ohraniceni = Side(border_style="thin", color="000000")
    for radek in aktivni_sesit[rozsah_bunek]:
        for bunka in radek:
            bunka.border = Border(top=styl_ohraniceni, bottom=styl_ohraniceni,
                                  left=styl_ohraniceni, right=styl_ohraniceni)

# vytvoreni prazdneho sesitu
sesit = Workbook()
# vytvoreni prazdnych listu
list1 = sesit.create_sheet("List 1", 1)
list2 = sesit.create_sheet("List 2", 0)

# seznam nazvu vsech listu v sesitu
print(sesit.sheetnames)
# smazani listu Sheet
del sesit["Sheet"]
# přejmenování listu
list2.title = "Data"

# iterace přes listy sešitu
for list_xls in sesit:
    print(list_xls.title)
# zápis do buňky A4
list2["A4"] = 7
list2["A5"] = "data"
# nová reference na list Data
data = sesit["Data"]
# přečtení hodnoty z buňky
print(data["A4"].value)
print(data["A1"].value)

# zapisování hodnot do buňek pomocí iterace přes řádky a sloupečky
for idx1, radek in enumerate(data.iter_rows(2, 4, max_col=3)):
    for idx2, bunka in enumerate(radek):
        bunka.value = f"radek {idx1}, sloupec {idx2}"

# spojení buňek A1:C1
data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
data["A1"] = CellRichText(
    TextBlock(InlineFont(b=True, sz=16, i=True), "Nadpis tabulky")
)
# nastavení ohraničení
nastav_ohraniceni(data, "A1:C5")
# šířka sloupce
data.column_dimensions["A"].width = 50
# smazání řádku
data.delete_rows(5, 1)
data.insert_rows(1, 3)
# uložit změny
sesit.save("sesit.xlsx")
sesit = load_workbook("sesit.xlsx")
# oblast spojených buňek
data = sesit["Data"]
print(data.merged_cells)
# rozpojení spojených buněk
data.unmerge_cells("A1:C1")
print("---------------")
print(f"{data.merged_cells} - spojene bunky ")
sesit.save("sesit.xlsx")



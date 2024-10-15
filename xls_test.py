from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import Border, Side


wb = Workbook()
ws = wb.active
print(ws)
ws1 = wb.create_sheet('Sheet1')
ws2 = wb.create_sheet('Sheet2')
ws2.title = "Nový"
ws2 = wb["Nový"]
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
ws2["A4"] = 7
hodnota = ws2["A4"]
print(f"hodnota {hodnota}")
for radek in ws2.iter_rows(min_row=1, max_col=3):
    for bunka in radek:
        print(bunka, bunka.value)
for idx1, radek in enumerate(ws2.iter_rows(min_row=1, max_col=3)):
    for idx2, bunka in enumerate(radek):
        bunka.value = f"{idx1},{idx2}"


ws2.merge_cells("E5:G5")
ws2["E5"] = "text"
ws2["E6"] = CellRichText(
    TextBlock(InlineFont(b=True, sz=24), "text tučně"), "další text"
)
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, bottom=thin, right=thin)
set_border(ws2, "A1:B5")
ws2.insert_rows(1, 1)
ws2.delete_rows(1, 1)
wb.save('test.xlsx')